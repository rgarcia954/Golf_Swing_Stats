import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def process_golf_stats(input_file, output_file, excluded_rows=None):
    """
    Process golf swing statistics CSV and generate XLSX with statistics.
    
    Parameters:
    - input_file: Path to input CSV file
    - output_file: Path to output XLSX file
    - excluded_rows: List of row numbers to exclude (1-indexed, matching the 'No.' column)
    """
    
    # Read the CSV file
    df = pd.read_csv(input_file)
    
    # Remove the existing AVG row if present
    df = df[df['No.'] != 'AVG'].copy()
    
    # Add an 'Include' column for tracking which rows to include
    df['Include'] = True
    
    # Exclude specified rows
    if excluded_rows:
        for row_num in excluded_rows:
            df.loc[df['No.'] == row_num, 'Include'] = False
    
    # Identify numeric columns (excluding No., Date, EQ, Include)
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    if 'No.' in numeric_cols:
        numeric_cols.remove('No.')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Golf Stats"
    
    # Write headers
    headers = ['No.', 'Date', 'EQ', 'Include'] + numeric_cols
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Write data rows
    row_idx = 2
    for _, row in df.iterrows():
        ws.cell(row=row_idx, column=1, value=row['No.'])
        ws.cell(row=row_idx, column=2, value=row['Date'])
        ws.cell(row=row_idx, column=3, value=row['EQ'])
        ws.cell(row=row_idx, column=4, value='Yes' if row['Include'] else 'No')
        
        for col_idx, col_name in enumerate(numeric_cols, 5):
            ws.cell(row=row_idx, column=col_idx, value=row[col_name])
        
        row_idx += 1
    
    # Add AVG row
    avg_row = row_idx
    ws.cell(row=avg_row, column=1, value='AVG')
    ws.cell(row=avg_row, column=2, value='')
    ws.cell(row=avg_row, column=3, value='')
    ws.cell(row=avg_row, column=4, value='')
    
    # Add formulas for averages
    for col_idx, col_name in enumerate(numeric_cols, 5):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # AVERAGEIF formula to only include rows where Include='Yes'
        formula = f'=AVERAGEIF($D$2:$D${avg_row-1},"Yes",{col_letter}2:{col_letter}{avg_row-1})'
        ws.cell(row=avg_row, column=col_idx, value=formula)
    
    # Style AVG row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=avg_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Add STDEV row
    stdev_row = avg_row + 1
    ws.cell(row=stdev_row, column=1, value='STDEV')
    ws.cell(row=stdev_row, column=2, value='')
    ws.cell(row=stdev_row, column=3, value='')
    ws.cell(row=stdev_row, column=4, value='')
    
    # Add formulas for standard deviation
    for col_idx, col_name in enumerate(numeric_cols, 5):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # Use SUMPRODUCT approach to avoid array formula issues with @
        # This calculates STDEV without needing array formula entry
        data_range = f'{col_letter}2:{col_letter}{avg_row-1}'
        include_range = f'$D$2:$D${avg_row-1}'
        avg_cell = f'{col_letter}{avg_row}'
        
        # Formula breakdown:
        # 1. COUNT included values: SUMPRODUCT(--($D$2:$D$9="Yes"))
        # 2. Sum of squared deviations: SUMPRODUCT((data-avg)^2,--($D$2:$D$9="Yes"))
        # 3. STDEV = SQRT(sum of squared deviations / (n-1))
        formula = (f'=SQRT(SUMPRODUCT((({data_range}-{avg_cell})^2),'
                   f'--({include_range}="Yes"))/'
                   f'(SUMPRODUCT(--({include_range}="Yes"))-1))')
        ws.cell(row=stdev_row, column=col_idx, value=formula)
    
    # Style STDEV row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=stdev_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save workbook
    wb.save(output_file)
    print(f"✓ Generated {output_file}")
    print(f"✓ Processed {len(df)} rows")
    print(f"✓ Excluded {sum(~df['Include'])} rows")
    print(f"✓ Included {sum(df['Include'])} rows in statistics")


# Example usage
if __name__ == "__main__":
    print("=" * 60)
    print("Golf Swing Statistics Analyzer")
    print("=" * 60)
    
    # Get input file name
    input_file = input("\nEnter input CSV file name (e.g., swingcaddie_6I.csv): ").strip()
    
    # Get output file name
    output_file = input("Enter output XLSX file name (e.g., golf_stats_analysis.xlsx): ").strip()
    
    # Ensure output file has .xlsx extension
    if not output_file.lower().endswith('.xlsx'):
        output_file += '.xlsx'
    
    # Get rows to exclude
    exclude_input = input("\nEnter row numbers to exclude, separated by commas (or press Enter for none): ").strip()
    
    excluded_rows = []
    if exclude_input:
        try:
            excluded_rows = [int(x.strip()) for x in exclude_input.split(',')]
            print(f"Will exclude rows: {excluded_rows}")
        except ValueError:
            print("Invalid input. No rows will be excluded.")
            excluded_rows = []
    
    print("\nProcessing...\n")
    
    try:
        process_golf_stats(
            input_file=input_file,
            output_file=output_file,
            excluded_rows=excluded_rows
        )
        print("\n" + "=" * 60)
        print("Processing complete!")
        print("=" * 60)
    except FileNotFoundError:
        print(f"\n❌ Error: Could not find file '{input_file}'")
        print("Please check the file name and try again.")
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")